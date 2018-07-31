'''
This file is for predict.
Load excel data and Predict using dead reckoning.
Output file is 'pred.xlsx'. 
'''


import openpyxl
from math import sin, cos

gravity = -9.79641227572363

# Excel file open
wb = openpyxl.load_workbook('data.xlsx')
wr = openpyxl.Workbook()
 
# Get Sheet of Excel
#ws = wb.active
wrs = wb.active

ri = 2
v = [0,0,0]
time_p = 0

wrs.cell(row=1, column=10).value = "pred - vx"
wrs.cell(row=1, column=11).value = "pred - vy"
wrs.cell(row=1, column=12).value = "pred - vz"

rows = wrs.rows
next(rows)

# Read Excel Data
for r in rows:
    if v[0]==0 and v[1]==0 and v[2]==0:
        v = [r[12].value, r[13].value, r[14].value]
        time_p = r[15].value
        ri += 1
        continue

    t = r[15].value - time_p
    time_p = r[15].value

    v = [v[0] + r[6].value*t, v[1] + r[7].value*t, v[2] + (r[8].value-gravity)*t]
    
    wrs.cell(row=ri, column=10).value = v[0]
    wrs.cell(row=ri, column=11).value = v[1]
    wrs.cell(row=ri, column=12).value = v[2]
    ri += 1
 

wb.save('pred-velo.xlsx')
print("End")
