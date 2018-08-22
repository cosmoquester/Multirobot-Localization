'''
This file is for predict.
Load excel data and Predict using dead reckoning.
Output file is 'pred.xlsx'. 
'''


import openpyxl
from math import sin, cos, pi
import numpy as np

gravity = -9.8

# Excel file open
wb = openpyxl.load_workbook('data.xlsx')
wr = openpyxl.Workbook()
 
# Get Sheet of Excel
#ws = wb.active
wrs = wb.active

ri = 2


wrs.cell(row=1, column=10).value = "pred - accX"
wrs.cell(row=1, column=11).value = "pred - accY"
wrs.cell(row=1, column=12).value = "pred - accZ"

rows = wrs.rows
next(rows)

# Read Excel Data
for r in rows:
    if not r[15].value:
        ri += 1
        continue


    roll = r[9].value
    fitch = r[10].value
    yaw = r[11].value


    alpha = roll
    beta = 2*pi - fitch
    gamma = yaw - pi/2

    basis = np.matrix([[0,1,0],[1,0,0],[0,0,1]])
    rotate_x = np.matrix([[1, 0, 0], [0, cos(alpha), -sin(alpha)], [0, sin(alpha), cos(alpha)]])
    rotate_y = np.matrix([[cos(beta), 0, sin(beta)], [0, 1, 0], [-sin(beta), 0, cos(beta)]])
    rotate_z = np.matrix([[cos(gamma), -sin(gamma), 0],[sin(gamma),cos(gamma),0],[0,0,1]])
    
    trans = rotate_x * rotate_y * rotate_z * basis
    #trans = np.matrix([[cos(gamma), cos(beta), 0],[sin(gamma), sin(beta), 0], [0,0,1]])
    
    acc = trans * np.matrix([[r[6].value], [r[7].value], [r[8].value + gravity]])
    
    wrs.cell(row=ri, column=10).value = acc.item(0)
    wrs.cell(row=ri, column=11).value = acc.item(1)
    wrs.cell(row=ri, column=12).value = acc.item(2)
    ri += 1
 

wb.save('pred-acc.xlsx')
print("End")
