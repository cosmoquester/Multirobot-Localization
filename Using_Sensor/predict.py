'''
This file is for predict.
Load excel data and Predict using dead reckoning.
Output file is 'pred.xlsx'. 
'''


import openpyxl
import matplotlib.pyplot as plt

gravity = -9.79641227572363

# Excel file open

wb = openpyxl.load_workbook('pred-velo.xlsx')
wr = openpyxl.Workbook()
 
# Get Sheet of Excel
ws = wb.active
wrs = wr.active
# ws = wb.get_sheet_by_name("Sheet1")


for i, val in enumerate(['Real-X', 'Real-Y', 'Real-Z', 'Estimate-X', 'Estimate-Y', 'Estimate-Z']):
    wrs.cell(row=1, column=i+1).value = val

time = 0
t=0
ok=False
Predict_position=[]
Previous_position=[]
orientation = [0,0,0]
errors=[]
ri=2

real_poss = [[],[],[]]
pred_poss = [[],[],[]]

# Read Excel Data
for r in ws.rows:
    if r[0].value==0 and r[1].value==0 and r[2].value==0 or r[0].value=='X-posi':
        continue
    
    if not ok:
        ok=True
        Previous_position=[r[0].value, r[1].value, r[2].value]
        Predict_position=Previous_position[:]
        time = r[15].value
        continue
    else:
        t = (r[15].value - time)/1000
        time = r[15].value
        Previous_position = [r[0].value, r[1].value, r[2].value]
        #orientation = [orientation[0] + r[9].value*t, orientation[1] + r[10].value*t, orientation[2] + r[11].value*t]

        # For showing Graph
        for ix in range(3):
            real_poss[ix].append(Previous_position[ix])
            pred_poss[ix].append(Predict_position[ix])

        # For Error Calculate
        errors.append(sum([(Predict_position[i]-Previous_position[i])**2 for i in range(3)])/3)
        for ci, value in enumerate(Previous_position + Predict_position) :
            wrs.cell(row=ri, column=ci+1).value = value
        ri+=1


    # Using "position = previous_position + velocity*time + 0.5*acceleration*time^2"
    #Predict_position = [Predict_position[0] + r[3].value*t, Predict_position[1] + r[4].value*t, Predict_position[2] + r[5].value*t]
    Predict_position = [Predict_position[0] + r[3].value*t + 0.5*r[9].value*t**2, Predict_position[1] + r[4].value*t+ 0.5*r[10].value*t**2, Predict_position[2] + r[5].value*t+ 0.5*r[11].value*t**2]

wr.save('pred.xlsx')
print("RMSD :",(sum(errors)/len(errors))**0.5)
plt.plot(real_poss[0], real_poss[1], color='yellow', label = 'Real')
plt.plot(pred_poss[0], pred_poss[1], color='violet', label = 'Predict')
plt.xlabel('x - velocity')
plt.ylabel('y - velcoity')
plt.legend(loc = 'upper right')
plt.show()
