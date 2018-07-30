'''
This file is for predict.
Load excel data and Predict using dead reckoning.
Output file is 'pred.xlsx'. 
'''


import openpyxl
from random import uniform as rand


NoiseLevel = 0.1

# Excel file open
wb = openpyxl.load_workbook('data.xlsx')
wr = openpyxl.Workbook()
 
# Get Sheet of Excel
ws = wb.active
wrs = wr.active
rows = ws.rows

# Write Names of Data
for i, c in enumerate(next(rows)):
    wrs.cell(row=1, column=i+1).value = c.value

# Read Excel Data
for i, r in enumerate(rows):
    for c in range(15):    
        wrs.cell(row=i+2, column = c+1).value = r[c].value + rand(-NoiseLevel, NoiseLevel)
    wrs.cell(row = i+2, column = 16).value = r[15].value


wr.save('data_noised.xlsx')
