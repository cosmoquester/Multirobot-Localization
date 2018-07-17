import openpyxl

gravity = -9.79641227572363

# 엑셀파일 열기
wb = openpyxl.load_workbook('data.xlsx')
wr = openpyxl.Workbook()
 
# 현재 Active Sheet 얻기
ws = wb.active
wrs = wr.active
# ws = wb.get_sheet_by_name("Sheet1")

wrs.cell(row=1, column=1).value = 'Real-X'
wrs.cell(row=1, column=2).value = 'Real-Y'
wrs.cell(row=1, column=3).value = 'Real-Z'
wrs.cell(row=1, column=4).value = 'Estimate-X'
wrs.cell(row=1, column=5).value = 'Estimate-Y'
wrs.cell(row=1, column=6).value = 'Estimate-Z'


t = 0.01
ok=False
Predict_position=[]
Previous_position=[]
errors=[]
ri=2

# Read Excel Data
for r in ws.rows:
    if r[0].value==0 and r[1].value==0 and r[2].value==0 or r[0].value=='X-posi':
        continue
    
    if not ok:
        ok=True
        Previous_position=[r[0].value, r[1].value, r[2].value]
        Predict_position=Previous_position[:]
        continue
    else:
        Previous_position = [r[0].value, r[1].value, r[2].value]
        errors.append(sum([(Predict_position[i]-Previous_position[i])**2 for i in range(3)])/3)
        for ci, value in enumerate(Previous_position+Predict_position) :
            wrs.cell(row=ri, column=ci+1).value = value
        ri+=1

    # Using "position = previous_position + velocity*time + 0.5*acceleration*time^2"
    Predict_position = [Predict_position[0] + r[12].value*t + r[6].value*0.5*t**2, Predict_position[1] + r[13].value*t + r[7].value*0.5*t**2, Predict_position[2] + r[14].value*t + (r[8].value-gravity)*0.5*t**2]

wr.save('pred.xlsx')
print("RMSD :",(sum(errors)/len(errors))**0.5)
